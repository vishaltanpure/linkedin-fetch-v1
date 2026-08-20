"""
Append a "Refreshed" sheet to a workbook using freshly scraped data from
output/batch-results.json. Maps our scraper's field names to the
workbook's exact Apify-style column names.

Usage:
    python scripts/refresh-xlsx.py "<path-to-xlsx>" "<path-to-batch-results.json>"
"""

import sys
import json
import openpyxl
from openpyxl.styles import Font

from xlsx_mapping import to_row, normalise_url


def main():
    if len(sys.argv) != 3:
        print("Usage: python scripts/refresh-xlsx.py <xlsx-path> <batch-results.json>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    json_path = sys.argv[2]

    with open(json_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    by_url = {normalise_url(r.get("profileUrl", "")): r for r in records}

    wb = openpyxl.load_workbook(xlsx_path)
    source = wb["Sheet1"]
    headers = [cell.value for cell in source[1]]

    sheet_name = "Refreshed"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    matched = 0
    missing = []

    for row_idx in range(2, source.max_row + 1):
        original_url = source.cell(row=row_idx, column=1).value
        record = by_url.get(normalise_url(original_url))

        if record is None:
            missing.append(original_url)
            ws.append([original_url] + [""] * (len(headers) - 1))
            continue

        matched += 1
        ws.append(to_row(record, headers))

    for col_idx, header in enumerate(headers, start=1):
        width = max(12, min(45, len(str(header)) + 4))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(xlsx_path)

    print(f"Matched {matched}/{source.max_row - 1} rows.")
    if missing:
        print("No fresh data found for:", missing)
    print(f"Saved sheet '{sheet_name}' into {xlsx_path}")


if __name__ == "__main__":
    main()
