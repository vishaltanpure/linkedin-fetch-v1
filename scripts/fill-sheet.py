"""
Fill in the remaining Apify-style columns (firstName, lastName, ...,
companyType) on an existing sheet that already has "originalQuery/query"
in column A, using freshly scraped data from output/batch-results.json.

Unlike refresh-xlsx.py (which adds a brand-new sheet), this writes into
the given sheet in place, matching rows by URL.

Usage:
    python scripts/fill-sheet.py "<path-to-xlsx>" "<path-to-batch-results.json>" "<sheet-name>"
"""

import sys
import json
import openpyxl

from xlsx_mapping import COLUMNS, to_row, normalise_url


def main():
    if len(sys.argv) != 4:
        print('Usage: python scripts/fill-sheet.py <xlsx-path> <batch-results.json> "<sheet-name>"')
        sys.exit(1)

    xlsx_path = sys.argv[1]
    json_path = sys.argv[2]
    sheet_name = sys.argv[3]

    with open(json_path, "r", encoding="utf-8") as f:
        records = json.load(f)

    by_url = {normalise_url(r.get("profileUrl", "")): r for r in records}

    wb = openpyxl.load_workbook(xlsx_path)

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]

    # Columns after A (originalQuery/query).
    fill_headers = COLUMNS[1:]
    for col_idx, header in enumerate(fill_headers, start=2):
        ws.cell(row=1, column=col_idx, value=header)

    matched = 0
    missing = []

    for row_idx in range(2, ws.max_row + 1):
        original_url = ws.cell(row=row_idx, column=1).value
        if not original_url:
            continue

        record = by_url.get(normalise_url(original_url))

        if record is None:
            missing.append(original_url)
            continue

        matched += 1
        row_values = to_row(record, COLUMNS)[1:]  # drop originalQuery/query, already in col A
        for col_idx, value in enumerate(row_values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(["originalQuery/query"] + fill_headers, start=1):
        width = max(12, min(45, len(str(header)) + 4))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(xlsx_path)

    print(f"Matched {matched}/{ws.max_row - 1} rows in '{sheet_name}'.")
    if missing:
        print("No fresh data found for:", missing)
    print(f"Saved into {xlsx_path}")


if __name__ == "__main__":
    main()
